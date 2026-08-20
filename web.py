#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
法拍监控系统 - Web 服务
- 关键词管理 / 标的列表 / 标的详情 / 实时日志 (SSE)
- 通过 /api/fp/run-now 触发抓取
- 启动 web.py 时**自动**启动内置调度器 + 立即跑一次全量抓取
- （scheduler.py 仍可独立运行，效果一致）
"""

import csv
import os
import queue
import threading
from datetime import datetime
from io import StringIO, BytesIO

from flask import Flask, render_template, request, jsonify, Response, redirect

import database as db
from config import (
    WEB_HOST, WEB_PORT, WEB_DEBUG, APP_NAME,
    BASE_DIR,
    SCHEDULE_DAILY_TIME,
    SCHEDULE_TIMEZONE,
)


app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

# ========== SSE 日志推送 ==========
log_queues = []
queues_lock = threading.Lock()


def send_log_to_frontend(log_text):
    """被 fp_scraper.py 调用的全局日志推送"""
    with queues_lock:
        for q in log_queues[:]:
            try:
                q.put_nowait(log_text)
            except Exception:
                pass


@app.route('/api/fp/logs/stream')
def fp_stream_logs():
    def generate():
        log_queue = queue.Queue()
        with queues_lock:
            log_queues.append(log_queue)
        try:
            yield "data: connected\n\n"
            while True:
                try:
                    line = log_queue.get(timeout=30)
                    yield f"data: {line}\n\n"
                except queue.Empty:
                    yield "data: heartbeat\n\n"
        except GeneratorExit:
            with queues_lock:
                if log_queue in log_queues:
                    log_queues.remove(log_queue)
    resp = Response(generate(), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    resp.headers['Connection'] = 'keep-alive'
    return resp


def log_sql(title, sql_str, **kwargs):
    lines = ["", "=" * 60, title]
    for k, v in kwargs.items():
        lines.append(f"   {k}: {v}")
    lines.append("-" * 60)
    for line in sql_str.strip().splitlines():
        lines.append("   " + line)
    lines.append("=" * 60)
    msg = "\n".join(lines)
    print(msg)
    send_log_to_frontend(msg)


# ========== 页面路由 ==========

@app.route('/')
def index():
    return redirect('/fp-items')


@app.route('/fp-keys')
def fp_keys():
    return render_template('fp_keys.html', app_name=APP_NAME, active_menu='fp-keys')


@app.route('/fp-items')
def fp_items():
    return render_template('fp_items.html', app_name=APP_NAME, active_menu='fp-items')


@app.route('/fp-item/<int:sf_item_id>')
def fp_item_detail(sf_item_id):
    return render_template('fp_item_detail.html',
                           app_name=APP_NAME, active_menu='fp-items',
                           sf_item_id=sf_item_id)


@app.route('/fp-logs')
def fp_logs():
    return render_template('fp_logs.html', app_name=APP_NAME, active_menu='fp-logs')


# ========== 首页汇总 ==========

@app.route('/api/fp/summary', methods=['GET'])
def fp_summary():
    return jsonify({'code': 0, 'msg': 'success', 'data': db.get_item_summary_stats()})


# ========== 关键词 API ==========

@app.route('/api/fp/keywords', methods=['GET'])
def fp_get_keywords():
    include_deleted = request.args.get('include_deleted', 'false').lower() == 'true'
    return jsonify({'code': 0, 'msg': 'success',
                    'data': db.get_all_keywords(include_deleted)})


@app.route('/api/fp/keywords-with-count', methods=['GET'])
def fp_get_keywords_with_count():
    keyword = request.args.get('keyword', '').strip()
    rows = db.get_items_with_keyword_count(keyword or None)
    for r in rows:
        r['item_count'] = int(r['item_count'] or 0)
    return jsonify({'code': 0, 'msg': 'success', 'data': rows})


@app.route('/api/fp/keyword', methods=['POST'])
def fp_add_keyword():
    data = request.get_json() or {}
    keyword = (data.get('keyword') or '').strip()
    remark = data.get('remark') or ''
    enabled = int(data.get('enabled', 1))
    if not keyword:
        return jsonify({'code': 1, 'msg': '关键词不能为空'})
    try:
        kid = db.add_keyword(keyword, remark, enabled)
        return jsonify({'code': 0, 'msg': '新增成功', 'data': {'id': kid}})
    except Exception as e:
        return jsonify({'code': 1, 'msg': f'新增失败: {e}'})


@app.route('/api/fp/keyword/<int:keyword_id>', methods=['PUT'])
def fp_update_keyword(keyword_id):
    data = request.get_json() or {}
    try:
        ok = db.update_keyword(
            keyword_id,
            keyword=data.get('keyword'),
            enabled=data.get('enabled'),
            remark=data.get('remark'),
        )
        return jsonify({'code': 0 if ok else 1, 'msg': '更新成功' if ok else '更新失败'})
    except Exception as e:
        return jsonify({'code': 1, 'msg': f'更新失败: {e}'})


@app.route('/api/fp/keyword/<int:keyword_id>/toggle', methods=['POST'])
def fp_toggle_keyword(keyword_id):
    data = request.get_json() or {}
    enabled = data.get('enabled')
    if enabled is None:
        return jsonify({'code': 1, 'msg': '参数错误'})
    try:
        ok = db.update_keyword(keyword_id, enabled=enabled)
        return jsonify({'code': 0 if ok else 1, 'msg': '操作成功' if ok else '操作失败',
                        'data': {'enabled': enabled}})
    except Exception as e:
        return jsonify({'code': 1, 'msg': f'操作失败: {e}'})


@app.route('/api/fp/keyword/<int:keyword_id>', methods=['DELETE'])
def fp_delete_keyword(keyword_id):
    try:
        ok = db.delete_keyword(keyword_id)
        return jsonify({'code': 0 if ok else 1, 'msg': '删除成功' if ok else '删除失败'})
    except Exception as e:
        return jsonify({'code': 1, 'msg': f'删除失败: {e}'})


# ========== 标的 API ==========

@app.route('/api/fp/items', methods=['GET'])
def fp_search_items():
    keyword = request.args.get('keyword', '').strip()
    status = request.args.get('status', '').strip()
    sf_item_id = request.args.get('sf_item_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    limit = request.args.get('limit', 200, type=int)
    offset = request.args.get('offset', 0, type=int)
    rows = db.search_items(
        keyword=keyword or None,
        status=status or None,
        sf_item_id=sf_item_id,
        date_from=date_from or None,
        date_to=date_to or None,
        limit=limit, offset=offset,
    )
    return jsonify({'code': 0, 'msg': 'success', 'data': rows})


@app.route('/api/fp/item/<int:sf_item_id>', methods=['GET'])
def fp_get_item(sf_item_id):
    item = db.get_item_by_sf_id(sf_item_id)
    history = db.get_item_history(sf_item_id, limit=60) if item else []
    return jsonify({'code': 0, 'msg': 'success', 'data': {'item': item, 'history': history}})


# ========== 标的导出 ==========

@app.route('/api/fp/export/items', methods=['GET'])
def fp_export_items():
    keyword = request.args.get('keyword', '').strip()
    status = request.args.get('status', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    rows = db.search_items(
        keyword=keyword or None,
        status=status or None,
        date_from=date_from or None,
        date_to=date_to or None,
        limit=5000,
    )
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(['标的ID', '标题', '状态', '起拍价', '当前价', '评估价', '报名数',
                     '竞价次数', '围观数', '开拍时间', '结束时间', '链接'])
    for r in rows:
        writer.writerow([
            r.get('sf_item_id'),
            r.get('title'),
            r.get('status'),
            f"{float(r['initial_price']):.2f}" if r.get('initial_price') is not None else '',
            f"{float(r['current_price']):.2f}" if r.get('current_price') is not None else '',
            f"{float(r['consult_price']):.2f}" if r.get('consult_price') is not None else '',
            r.get('apply_count'),
            r.get('bid_count'),
            r.get('viewer_count'),
            r['start_at'].strftime('%Y-%m-%d %H:%M:%S') if r.get('start_at') else '',
            r['end_at'].strftime('%Y-%m-%d %H:%M:%S') if r.get('end_at') else '',
            r.get('item_url'),
        ])
    filename = f"fp_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    resp = Response(buf.getvalue(), mimetype='text/csv; charset=utf-8')
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@app.route('/api/fp/export/items-xlsx', methods=['GET'])
def fp_export_items_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension

    keyword = request.args.get('keyword', '').strip()
    status = request.args.get('status', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    rows = db.search_items(
        keyword=keyword or None,
        status=status or None,
        date_from=date_from or None,
        date_to=date_to or None,
        limit=5000,
    )
    columns = [
        {'name': '标的ID', 'width': 16, 'align': 'center', 'wrap': False},
        {'name': '标题', 'width': 50, 'align': 'left', 'wrap': True},
        {'name': '状态', 'width': 8, 'align': 'center', 'wrap': False},
        {'name': '起拍价', 'width': 12, 'align': 'right', 'wrap': False},
        {'name': '当前价', 'width': 12, 'align': 'right', 'wrap': False},
        {'name': '评估价', 'width': 12, 'align': 'right', 'wrap': False},
        {'name': '报名', 'width': 8, 'align': 'center', 'wrap': False},
        {'name': '竞价', 'width': 8, 'align': 'center', 'wrap': False},
        {'name': '围观', 'width': 10, 'align': 'center', 'wrap': False},
        {'name': '开拍时间', 'width': 18, 'align': 'center', 'wrap': False},
        {'name': '结束时间', 'width': 18, 'align': 'center', 'wrap': False},
        {'name': '链接', 'width': 40, 'align': 'left', 'wrap': True},
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "标的列表"
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        letter = get_column_letter(col_idx)
        cd = ColumnDimension(ws, min=col_idx, max=col_idx, width=col['width'])
        ws.column_dimensions[letter] = cd
    body_font = Font(name='微软雅黑', size=10)
    for r_idx, r in enumerate(rows, 2):
        values = [
            r.get('sf_item_id'),
            r.get('title'),
            r.get('status'),
            float(r['initial_price']) if r.get('initial_price') is not None else '',
            float(r['current_price']) if r.get('current_price') is not None else '',
            float(r['consult_price']) if r.get('consult_price') is not None else '',
            r.get('apply_count'),
            r.get('bid_count'),
            r.get('viewer_count'),
            r['start_at'].strftime('%Y-%m-%d %H:%M:%S') if r.get('start_at') else '',
            r['end_at'].strftime('%Y-%m-%d %H:%M:%S') if r.get('end_at') else '',
            r.get('item_url'),
        ]
        for c_idx, val in enumerate(values, 1):
            col = columns[c_idx - 1]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal=col['align'], vertical='top', wrap_text=col['wrap']
            )
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"fp_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp.headers['Content-Length'] = str(len(bio.getvalue()))
    return resp


# ========== 运行日志 API ==========

@app.route('/api/fp/run-logs', methods=['GET'])
def fp_run_logs():
    limit = request.args.get('limit', 30, type=int)
    return jsonify({'code': 0, 'msg': 'success',
                    'data': db.get_recent_run_logs(limit)})


# ========== 触发抓取 ==========

@app.route('/api/fp/run-now', methods=['POST'])
def fp_run_now():
    """触发一次抓取（在前端按钮点击时调用）"""
    data = request.get_json() or {}
    keyword = (data.get('keyword') or '').strip()
    keyword_id = data.get('keyword_id')

    def _run():
        try:
            send_log_to_frontend("=" * 50)
            send_log_to_frontend(f"📌 手动触发抓取: keyword={keyword or '-'} keyword_id={keyword_id or '-'}")
            send_log_to_frontend("=" * 50)
            import fp_scraper
            if keyword_id:
                row = db.get_keyword_by_id(int(keyword_id))
                if row:
                    fp_scraper.scrape_keyword(row)
                else:
                    send_log_to_frontend(f"❌ 关键词 id={keyword_id} 不存在")
            elif keyword:
                fp_scraper.scrape_keyword(keyword)
            else:
                fp_scraper.scrape_all_enabled()
        except Exception as e:
            send_log_to_frontend(f"❌ 抓取异常: {e}")

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'code': 0, 'msg': '已触发抓取任务'})


# ========== 调试：抓取指定 URL 并打印原始数据（不入库） ==========

@app.route('/api/fp/debug-fetch', methods=['POST'])
def fp_debug_fetch():
    """
    调试用：抓取指定完整 URL，把响应大小 + items 条数 + 前 20 条原始数据返回
    body: {"url": "https://sf.taobao.com/item_list.htm?..."}
    """
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'code': 1, 'msg': 'url 不能为空'})

    def _run():
        import re, json
        import subprocess
        try:
            send_log_to_frontend("=" * 50)
            send_log_to_frontend(f"🔧 DEBUG FETCH: {url}")
            send_log_to_frontend("=" * 50)
            proc = subprocess.run(
                ['curl', '-sL', '--max-time', '30',
                 '-A', 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                 '-H', 'Referer: https://sf.taobao.com/',
                 url],
                capture_output=True, timeout=60
            )
            raw = proc.stdout
            html = raw.decode('gbk', errors='ignore')
            send_log_to_frontend(f"📦 响应字节数: {len(raw)}")

            # 检查 charset
            cm = re.search(r'<meta[^>]*charset=[\"]?([^\"]+)[\"]?', html, re.IGNORECASE)
            if cm:
                send_log_to_frontend(f"🔤 charset: {cm.group(1)}")

            # 检查验证码
            if '_____tmd_____' in html or 'action":"captcha"' in html:
                send_log_to_frontend("❌ 触发验证码拦截！")
                return

            # 抓 sf-item-list-data
            m = re.search(r'sf-item-list-data[^>]*>(.*?)</script>', html, re.S)
            if not m:
                send_log_to_frontend("⚠️ 没找到 sf-item-list-data")
                send_log_to_frontend(f"HTML 前 300 字符: {html[:300]}")
                return

            j = json.loads(m.group(1))
            items = j.get('data', [])
            send_log_to_frontend(f"📊 items 条数: {len(items)}")
            if not items:
                send_log_to_frontend("⚠️ 该 URL 返回的 items 为空（淘宝搜索 API 对该 q 无结果）")
                return

            # 打印所有
            st = {}
            for it in items:
                s = it.get('status')
                st[s] = st.get(s, 0) + 1
            send_log_to_frontend(f"状态分布: {st}")
            send_log_to_frontend(f"{'─'*60}")
            send_log_to_frontend(f"{'序号':<4} {'标的ID':<16} {'状态':<8} {'当前价':<14} {'标题'}")
            send_log_to_frontend(f"{'─'*60}")
            for i, it in enumerate(items, 1):
                sid = it.get('id') or '-'
                s = it.get('status') or '-'
                price = it.get('currentPrice') or 0
                title = (it.get('title') or '').replace('\n', ' ').strip()
                price_str = f"{float(price):,.2f}" if price else '-'
                send_log_to_frontend(f"{i:<4} {str(sid):<16} {s:<8} {price_str:<14} {title[:60]}")
        except Exception as e:
            send_log_to_frontend(f"❌ DEBUG FETCH 异常: {e}")

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'code': 0, 'msg': '已触发调试抓取'})


# ========== 启动 ==========

# ========== 启动 ==========

def _start_embedded_scheduler():
    """
    启动 web 时自动启动内置调度器：
    - 每天 SCHEDULE_DAILY_TIME 自动跑
    - 启动时会打印下次触发时间（cron 不会触发当天已过的时间）
    """
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        import threading

        # 从 config 读取定时参数
        hour, minute = 9, 0
        try:
            h, m = SCHEDULE_DAILY_TIME.strip().split(':')
            hour, minute = int(h), int(m)
        except Exception:
            pass
        tz = SCHEDULE_TIMEZONE

        sched = BackgroundScheduler(timezone=tz)
        sched.add_job(
            lambda: threading.Thread(target=_run_daily_scrape, daemon=True).start(),
            CronTrigger(hour=hour, minute=minute),
            id='daily_scrape',
            name=f'每日{SCHEDULE_DAILY_TIME}抓取',
            max_instances=1,
            coalesce=True,
            misfire_grace_time=600,
        )
        sched.start()
        # 打印下次触发时间（关键：cron 不会触发已过的时间）
        job = sched.get_job('daily_scrape')
        next_run = job.next_run_time if job else None
        msg = f"⏰ 内置调度器已启动：每日 {hour:02d}:{minute:02d} 抓取任务"
        if next_run:
            msg += f"\n   ⏱️  下次触发时间: {next_run.strftime('%Y-%m-%d %H:%M:%S %Z')}"
        from datetime import datetime as _dt
        now = _dt.now()
        if next_run and next_run.date() > now.date():
            msg += f"\n   ⚠️ 注意：今天的 {hour:02d}:{minute:02d} 已过，下次要等明天！"
        send_log_to_frontend(msg)
        print(msg.replace('\n', ' | '))
        return sched
    except Exception as e:
        send_log_to_frontend(f"⚠️ 调度器启动失败: {e}")
        return None


def _run_daily_scrape():
    """每日定时抓取任务（通过 SSE 推送到日志）"""
    try:
        send_log_to_frontend("=" * 50)
        send_log_to_frontend(f"⏰ 触发每日定时抓取 @ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        send_log_to_frontend("=" * 50)
        import fp_scraper
        results = fp_scraper.scrape_all_enabled()
        ok = sum(1 for r in results if r.get('status') in ('success', 'stopped'))
        fail = len(results) - ok
        send_log_to_frontend(f"✅ 每日抓取完成: 成功/停止 {ok} / 失败 {fail} / 共 {len(results)}")
    except Exception as e:
        send_log_to_frontend(f"❌ 每日抓取异常: {e}")


# ========== 启动时自动跑批（模块级） ==========
# web 模块加载时自动启动内置调度器；每天 SCHEDULE_DAILY_TIME 自动跑一次
# 启动 web 时**不会**立刻抓数据（除非手动点击前端"立即抓取"按钮）
_WEB_AUTO_STARTED = False


def _bootstrap_auto_run():
    """web 模块加载时自动启动内置调度器"""
    global _WEB_AUTO_STARTED
    if _WEB_AUTO_STARTED:
        return
    _WEB_AUTO_STARTED = True

    # 启动内置调度器（每日 SCHEDULE_DAILY_TIME 自动跑批）
    _start_embedded_scheduler()
    print(f"⏰ 内置调度器已启动：每日 {SCHEDULE_DAILY_TIME} 抓取任务")
    print(f"💡 如需立即抓取，请点击前端页面上的「立即抓取」按钮或调用 /api/fp/run-now")


_bootstrap_auto_run()


if __name__ == '__main__':
    print(f"=== {APP_NAME} 启动 ===")
    print(f"Web:  http://127.0.0.1:{WEB_PORT}")
    print(f"⏰ 内置调度器已启动：每日 {SCHEDULE_DAILY_TIME} 抓取任务")
    print("💡 启动时不会立即抓取，到点才会跑")
    print("（如需独立运行调度器: python scheduler.py）")
    app.run(host=WEB_HOST, port=WEB_PORT, debug=WEB_DEBUG, threaded=True)
